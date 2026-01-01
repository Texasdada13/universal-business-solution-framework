"""
Report Generator - Universal Business Solution Framework

Multi-format report generation supporting CSV, JSON, Excel, and structured reports.
Provides a unified interface for exporting data to various formats.

Usage:
```python
from patterns.export import ReportGenerator, ReportFormat, ReportSection

# Create generator
generator = ReportGenerator(title="Q4 Analysis Report")

# Add sections
generator.add_section(ReportSection(
    title="Executive Summary",
    content=summary_data,
    section_type="summary"
))
generator.add_section(ReportSection(
    title="Detailed Analysis",
    content=detail_df,
    section_type="table"
))

# Export to multiple formats
generator.export('report.xlsx', ReportFormat.EXCEL)
generator.export('report.json', ReportFormat.JSON)
generator.export('report.csv', ReportFormat.CSV)
```
"""

from dataclasses import dataclass, field
from enum import Enum
from typing import Dict, List, Optional, Any, Union, Callable
from pathlib import Path
from datetime import datetime
import json
import csv
import io


class ReportFormat(Enum):
    """Supported report formats"""
    CSV = "csv"
    JSON = "json"
    EXCEL = "excel"
    HTML = "html"
    MARKDOWN = "markdown"


class SectionType(Enum):
    """Types of report sections"""
    SUMMARY = "summary"
    TABLE = "table"
    METRICS = "metrics"
    CHART_DATA = "chart_data"
    TEXT = "text"
    LIST = "list"
    KEY_VALUE = "key_value"


@dataclass
class ReportSection:
    """A section of the report"""
    title: str
    content: Any
    section_type: Union[SectionType, str] = SectionType.TEXT
    description: str = ""
    metadata: Dict[str, Any] = field(default_factory=dict)

    def __post_init__(self):
        if isinstance(self.section_type, str):
            try:
                self.section_type = SectionType(self.section_type)
            except ValueError:
                self.section_type = SectionType.TEXT


@dataclass
class ReportMetadata:
    """Report metadata"""
    title: str
    subtitle: str = ""
    author: str = ""
    created_at: datetime = field(default_factory=datetime.now)
    version: str = "1.0"
    tags: List[str] = field(default_factory=list)
    custom: Dict[str, Any] = field(default_factory=dict)

    def to_dict(self) -> Dict:
        return {
            'title': self.title,
            'subtitle': self.subtitle,
            'author': self.author,
            'created_at': self.created_at.isoformat(),
            'version': self.version,
            'tags': self.tags,
            **self.custom
        }


class ReportGenerator:
    """
    Multi-format report generator.

    Supports structured reports with multiple sections,
    exportable to various formats.

    Example:
    ```python
    report = ReportGenerator(
        title="Monthly Performance Report",
        subtitle="December 2024"
    )

    # Add summary metrics
    report.add_metrics("Key Metrics", {
        "Total Revenue": "$1.2M",
        "Growth Rate": "15%",
        "Active Users": "50,000"
    })

    # Add data table
    report.add_table("Sales by Region", sales_data)

    # Export
    report.export("report.xlsx", ReportFormat.EXCEL)
    ```
    """

    def __init__(
        self,
        title: str,
        subtitle: str = "",
        author: str = "",
        include_timestamp: bool = True
    ):
        """
        Initialize report generator.

        Args:
            title: Report title
            subtitle: Report subtitle
            author: Report author
            include_timestamp: Include generation timestamp
        """
        self.metadata = ReportMetadata(
            title=title,
            subtitle=subtitle,
            author=author
        )
        self.sections: List[ReportSection] = []
        self.include_timestamp = include_timestamp
        self._export_handlers: Dict[ReportFormat, Callable] = {
            ReportFormat.CSV: self._export_csv,
            ReportFormat.JSON: self._export_json,
            ReportFormat.EXCEL: self._export_excel,
            ReportFormat.HTML: self._export_html,
            ReportFormat.MARKDOWN: self._export_markdown,
        }

    def add_section(self, section: ReportSection) -> 'ReportGenerator':
        """Add a section to the report"""
        self.sections.append(section)
        return self

    def add_text(self, title: str, content: str, description: str = "") -> 'ReportGenerator':
        """Add a text section"""
        return self.add_section(ReportSection(
            title=title,
            content=content,
            section_type=SectionType.TEXT,
            description=description
        ))

    def add_metrics(self, title: str, metrics: Dict[str, Any], description: str = "") -> 'ReportGenerator':
        """Add a metrics section (key-value pairs)"""
        return self.add_section(ReportSection(
            title=title,
            content=metrics,
            section_type=SectionType.METRICS,
            description=description
        ))

    def add_table(
        self,
        title: str,
        data: Union[List[Dict], 'pd.DataFrame'],
        description: str = ""
    ) -> 'ReportGenerator':
        """Add a table section"""
        # Convert DataFrame to list of dicts if needed
        if hasattr(data, 'to_dict'):
            data = data.to_dict('records')

        return self.add_section(ReportSection(
            title=title,
            content=data,
            section_type=SectionType.TABLE,
            description=description
        ))

    def add_list(self, title: str, items: List[str], description: str = "") -> 'ReportGenerator':
        """Add a list section"""
        return self.add_section(ReportSection(
            title=title,
            content=items,
            section_type=SectionType.LIST,
            description=description
        ))

    def add_summary(self, title: str, content: Dict[str, Any], description: str = "") -> 'ReportGenerator':
        """Add a summary section"""
        return self.add_section(ReportSection(
            title=title,
            content=content,
            section_type=SectionType.SUMMARY,
            description=description
        ))

    def export(
        self,
        path: Union[str, Path],
        format: ReportFormat = ReportFormat.JSON
    ) -> Path:
        """
        Export report to file.

        Args:
            path: Output file path
            format: Output format

        Returns:
            Path to exported file
        """
        path = Path(path)

        if format not in self._export_handlers:
            raise ValueError(f"Unsupported format: {format}")

        handler = self._export_handlers[format]
        handler(path)

        return path

    def export_all(
        self,
        base_path: Union[str, Path],
        formats: Optional[List[ReportFormat]] = None
    ) -> Dict[ReportFormat, Path]:
        """
        Export to multiple formats.

        Args:
            base_path: Base path without extension
            formats: Formats to export (default: all)

        Returns:
            Dict mapping format to exported path
        """
        base_path = Path(base_path)
        if formats is None:
            formats = list(ReportFormat)

        results = {}
        for fmt in formats:
            ext = self._get_extension(fmt)
            path = base_path.with_suffix(ext)
            try:
                self.export(path, fmt)
                results[fmt] = path
            except Exception as e:
                print(f"Warning: Could not export to {fmt.value}: {e}")

        return results

    def to_dict(self) -> Dict:
        """Convert report to dictionary"""
        return {
            'metadata': self.metadata.to_dict(),
            'sections': [
                {
                    'title': s.title,
                    'type': s.section_type.value if isinstance(s.section_type, SectionType) else s.section_type,
                    'description': s.description,
                    'content': s.content,
                    'metadata': s.metadata
                }
                for s in self.sections
            ],
            'generated_at': datetime.now().isoformat() if self.include_timestamp else None
        }

    def _get_extension(self, format: ReportFormat) -> str:
        """Get file extension for format"""
        extensions = {
            ReportFormat.CSV: '.csv',
            ReportFormat.JSON: '.json',
            ReportFormat.EXCEL: '.xlsx',
            ReportFormat.HTML: '.html',
            ReportFormat.MARKDOWN: '.md',
        }
        return extensions.get(format, '.txt')

    # ==================== Export Handlers ====================

    def _export_json(self, path: Path) -> None:
        """Export to JSON"""
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(self.to_dict(), f, indent=2, default=str)

    def _export_csv(self, path: Path) -> None:
        """Export to CSV (flattened format)"""
        rows = []

        # Add metadata
        rows.append(['Report Title', self.metadata.title])
        rows.append(['Subtitle', self.metadata.subtitle])
        rows.append(['Generated', datetime.now().isoformat()])
        rows.append([])  # Empty row

        # Add sections
        for section in self.sections:
            rows.append(['Section', section.title])
            rows.append(['Type', section.section_type.value if isinstance(section.section_type, SectionType) else section.section_type])

            if section.section_type == SectionType.TABLE and isinstance(section.content, list) and section.content:
                # Write table header
                headers = list(section.content[0].keys())
                rows.append(headers)
                # Write table rows
                for item in section.content:
                    rows.append([item.get(h, '') for h in headers])

            elif section.section_type == SectionType.METRICS and isinstance(section.content, dict):
                for key, value in section.content.items():
                    rows.append([key, value])

            elif section.section_type == SectionType.LIST and isinstance(section.content, list):
                for item in section.content:
                    rows.append([item])

            else:
                rows.append([str(section.content)])

            rows.append([])  # Empty row between sections

        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(rows)

    def _export_excel(self, path: Path) -> None:
        """Export to Excel with multiple sheets"""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
        except ImportError:
            # Fallback to CSV if openpyxl not available
            print("Warning: openpyxl not available, falling back to CSV")
            self._export_csv(path.with_suffix('.csv'))
            return

        wb = Workbook()

        # Style definitions
        header_font = Font(bold=True, color='FFFFFF', size=12)
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"

        # Add title
        ws_summary['A1'] = self.metadata.title
        ws_summary['A1'].font = Font(bold=True, size=16)
        if self.metadata.subtitle:
            ws_summary['A2'] = self.metadata.subtitle
            ws_summary['A2'].font = Font(italic=True, size=12)

        row = 4
        ws_summary[f'A{row}'] = 'Generated:'
        ws_summary[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M')

        # Add metrics sections to summary
        row += 2
        for section in self.sections:
            if section.section_type == SectionType.METRICS and isinstance(section.content, dict):
                ws_summary[f'A{row}'] = section.title
                ws_summary[f'A{row}'].font = Font(bold=True, size=11)
                row += 1
                for key, value in section.content.items():
                    ws_summary[f'A{row}'] = key
                    ws_summary[f'B{row}'] = value
                    row += 1
                row += 1

        # Auto-adjust column widths for summary
        ws_summary.column_dimensions['A'].width = 30
        ws_summary.column_dimensions['B'].width = 25

        # Add table sections as separate sheets
        for section in self.sections:
            if section.section_type == SectionType.TABLE and isinstance(section.content, list) and section.content:
                # Create sheet with safe name
                sheet_name = section.title[:31]  # Excel limit
                sheet_name = ''.join(c for c in sheet_name if c not in '[]:*?/\\')
                ws = wb.create_sheet(title=sheet_name)

                # Convert to DataFrame
                df = pd.DataFrame(section.content)

                # Write header
                for col_idx, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=column)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = thin_border

                # Write data
                for row_idx, row_data in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.border = thin_border

                # Auto-adjust column widths
                for col_idx, column in enumerate(df.columns, 1):
                    max_length = max(
                        len(str(column)),
                        df[column].astype(str).str.len().max() if len(df) > 0 else 0
                    )
                    ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = min(max_length + 2, 50)

        wb.save(path)

    def _export_html(self, path: Path) -> None:
        """Export to HTML"""
        html_parts = [
            '<!DOCTYPE html>',
            '<html>',
            '<head>',
            f'<title>{self.metadata.title}</title>',
            '<style>',
            'body { font-family: Arial, sans-serif; margin: 40px; }',
            'h1 { color: #2c3e50; }',
            'h2 { color: #34495e; border-bottom: 2px solid #3498db; padding-bottom: 10px; }',
            'table { border-collapse: collapse; width: 100%; margin: 20px 0; }',
            'th, td { border: 1px solid #ddd; padding: 12px; text-align: left; }',
            'th { background-color: #3498db; color: white; }',
            'tr:nth-child(even) { background-color: #f2f2f2; }',
            '.metrics { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; }',
            '.metric-card { background: #f8f9fa; padding: 20px; border-radius: 8px; border-left: 4px solid #3498db; }',
            '.metric-value { font-size: 24px; font-weight: bold; color: #2c3e50; }',
            '.metric-label { color: #7f8c8d; margin-top: 5px; }',
            '.timestamp { color: #95a5a6; font-size: 12px; }',
            '</style>',
            '</head>',
            '<body>',
            f'<h1>{self.metadata.title}</h1>',
        ]

        if self.metadata.subtitle:
            html_parts.append(f'<p><em>{self.metadata.subtitle}</em></p>')

        if self.include_timestamp:
            html_parts.append(f'<p class="timestamp">Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}</p>')

        for section in self.sections:
            html_parts.append(f'<h2>{section.title}</h2>')

            if section.description:
                html_parts.append(f'<p>{section.description}</p>')

            if section.section_type == SectionType.METRICS and isinstance(section.content, dict):
                html_parts.append('<div class="metrics">')
                for key, value in section.content.items():
                    html_parts.append(f'<div class="metric-card">')
                    html_parts.append(f'<div class="metric-value">{value}</div>')
                    html_parts.append(f'<div class="metric-label">{key}</div>')
                    html_parts.append('</div>')
                html_parts.append('</div>')

            elif section.section_type == SectionType.TABLE and isinstance(section.content, list) and section.content:
                html_parts.append('<table>')
                # Header
                headers = list(section.content[0].keys())
                html_parts.append('<tr>' + ''.join(f'<th>{h}</th>' for h in headers) + '</tr>')
                # Rows
                for item in section.content:
                    cells = ''.join(f'<td>{item.get(h, "")}</td>' for h in headers)
                    html_parts.append(f'<tr>{cells}</tr>')
                html_parts.append('</table>')

            elif section.section_type == SectionType.LIST and isinstance(section.content, list):
                html_parts.append('<ul>')
                for item in section.content:
                    html_parts.append(f'<li>{item}</li>')
                html_parts.append('</ul>')

            else:
                html_parts.append(f'<p>{section.content}</p>')

        html_parts.extend(['</body>', '</html>'])

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(html_parts))

    def _export_markdown(self, path: Path) -> None:
        """Export to Markdown"""
        lines = [
            f'# {self.metadata.title}',
            ''
        ]

        if self.metadata.subtitle:
            lines.append(f'*{self.metadata.subtitle}*')
            lines.append('')

        if self.include_timestamp:
            lines.append(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}')
            lines.append('')

        lines.append('---')
        lines.append('')

        for section in self.sections:
            lines.append(f'## {section.title}')
            lines.append('')

            if section.description:
                lines.append(section.description)
                lines.append('')

            if section.section_type == SectionType.METRICS and isinstance(section.content, dict):
                for key, value in section.content.items():
                    lines.append(f'- **{key}**: {value}')
                lines.append('')

            elif section.section_type == SectionType.TABLE and isinstance(section.content, list) and section.content:
                headers = list(section.content[0].keys())
                # Header row
                lines.append('| ' + ' | '.join(headers) + ' |')
                # Separator
                lines.append('| ' + ' | '.join(['---'] * len(headers)) + ' |')
                # Data rows
                for item in section.content:
                    row = '| ' + ' | '.join(str(item.get(h, '')) for h in headers) + ' |'
                    lines.append(row)
                lines.append('')

            elif section.section_type == SectionType.LIST and isinstance(section.content, list):
                for item in section.content:
                    lines.append(f'- {item}')
                lines.append('')

            else:
                lines.append(str(section.content))
                lines.append('')

        with open(path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(lines))


class DataExporter:
    """
    Simple data exporter for quick exports.

    Example:
    ```python
    exporter = DataExporter()
    exporter.to_csv(data, 'output.csv')
    exporter.to_json(data, 'output.json')
    exporter.to_excel(data, 'output.xlsx')
    ```
    """

    def __init__(self, include_timestamp: bool = False):
        """
        Initialize exporter.

        Args:
            include_timestamp: Add timestamp to filename
        """
        self.include_timestamp = include_timestamp

    def _get_path(self, path: Union[str, Path], extension: str) -> Path:
        """Get path with optional timestamp"""
        path = Path(path)
        if self.include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            return path.parent / f"{path.stem}_{timestamp}{extension}"
        return path

    def to_csv(
        self,
        data: Union[List[Dict], 'pd.DataFrame'],
        path: Union[str, Path],
        **kwargs
    ) -> Path:
        """Export data to CSV"""
        path = self._get_path(path, '.csv')

        if hasattr(data, 'to_csv'):
            data.to_csv(path, index=False, **kwargs)
        else:
            with open(path, 'w', newline='', encoding='utf-8') as f:
                if data:
                    writer = csv.DictWriter(f, fieldnames=data[0].keys())
                    writer.writeheader()
                    writer.writerows(data)

        return path

    def to_json(
        self,
        data: Any,
        path: Union[str, Path],
        indent: int = 2,
        **kwargs
    ) -> Path:
        """Export data to JSON"""
        path = self._get_path(path, '.json')

        if hasattr(data, 'to_dict'):
            data = data.to_dict('records')

        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, default=str, **kwargs)

        return path

    def to_excel(
        self,
        data: Union[List[Dict], 'pd.DataFrame', Dict[str, Any]],
        path: Union[str, Path],
        sheet_name: str = 'Data',
        **kwargs
    ) -> Path:
        """Export data to Excel"""
        path = self._get_path(path, '.xlsx')

        try:
            import pandas as pd

            if isinstance(data, dict) and all(isinstance(v, (list, pd.DataFrame)) for v in data.values()):
                # Multiple sheets
                with pd.ExcelWriter(path, engine='openpyxl') as writer:
                    for sheet, sheet_data in data.items():
                        df = pd.DataFrame(sheet_data) if not isinstance(sheet_data, pd.DataFrame) else sheet_data
                        df.to_excel(writer, sheet_name=sheet[:31], index=False)
            else:
                # Single sheet
                df = pd.DataFrame(data) if not isinstance(data, pd.DataFrame) else data
                df.to_excel(path, sheet_name=sheet_name, index=False, **kwargs)

        except ImportError:
            # Fallback to CSV
            print("Warning: pandas/openpyxl not available, falling back to CSV")
            return self.to_csv(data, path.with_suffix('.csv'))

        return path


# ==================== Factory Functions ====================


def create_simple_report(title: str, data: List[Dict]) -> ReportGenerator:
    """Create a simple report with one data table"""
    report = ReportGenerator(title=title)
    report.add_table("Data", data)
    return report


def create_analysis_report(
    title: str,
    metrics: Dict[str, Any],
    data: List[Dict],
    recommendations: Optional[List[str]] = None
) -> ReportGenerator:
    """
    Create a standard analysis report.

    Includes: Metrics, Data Table, Recommendations
    """
    report = ReportGenerator(title=title)
    report.add_metrics("Key Metrics", metrics)
    report.add_table("Detailed Data", data)
    if recommendations:
        report.add_list("Recommendations", recommendations)
    return report


def create_executive_summary(
    title: str,
    overview: str,
    highlights: List[str],
    metrics: Dict[str, Any],
    next_steps: Optional[List[str]] = None
) -> ReportGenerator:
    """Create an executive summary report"""
    report = ReportGenerator(title=title, subtitle="Executive Summary")
    report.add_text("Overview", overview)
    report.add_list("Key Highlights", highlights)
    report.add_metrics("Performance Metrics", metrics)
    if next_steps:
        report.add_list("Next Steps", next_steps)
    return report


# ==================== Exports ====================

__all__ = [
    'ReportGenerator',
    'ReportFormat',
    'ReportSection',
    'ReportMetadata',
    'SectionType',
    'DataExporter',
    'create_simple_report',
    'create_analysis_report',
    'create_executive_summary',
]
